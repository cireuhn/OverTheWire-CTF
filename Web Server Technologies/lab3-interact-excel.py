from openpyxl import load_workbook   
import xlwt

# #2 Using Python read the Excel spreadsheet you created.
excel_file = r'Web Server Technologies\Lab3-excel.xlsx'

wb = load_workbook(excel_file)
ws = wb.active

cell_value = ws['A1'].value
print(cell_value)

# #3 Using Python code modify the Excel spreadsheet used in #1 and #2 by changing your name to Alice.
ws['A1'] = 'Alice'
#wb.save(excel_file)

new_cell_value = ws['A1'].value
print(new_cell_value)
# Submit a screenshot of your code and the results.